import streamlit as st
import pandas as pd
import pyodbc
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import plotly.express as px
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="📦 Channel Despatch Summary", layout="wide")
st.title("🚚 Daily Despatch Summary")

# ------------------ DB CONNECT ------------------
def connect_db():
    try:
        return pyodbc.connect(
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=mptcecommerce-sql-server.database.windows.net;"
            "DATABASE=mptcecommerce-db;"
            "UID=mptcadmin;"
            "PWD=Mptc@2025;"
            "Connection Timeout=30"
        )
    except Exception as e:
        st.error(f"❌ Database connection failed: {e}")
        return None

# ------------------ DATE FILTER UTILITY ------------------
def get_range_from_option(option, available_dates):
    if not available_dates:
        return None, None

    latest_date = max(available_dates)

    if option == "Yesterday":
        # Always use the latest available date
        return latest_date, latest_date
    elif option == "Last 7 Days":
        return latest_date - timedelta(days=6), latest_date
    elif option == "Last 30 Days":
        return latest_date - timedelta(days=29), latest_date
    elif option == "Last 3 Months":
        return latest_date - relativedelta(months=3), latest_date
    elif option == "Last 6 Months":
        return latest_date - relativedelta(months=6), latest_date
    elif option == "Last 12 Months":
        return latest_date - relativedelta(months=12), latest_date
    return None, None

# ------------------ DATE FILTER UI ------------------
st.sidebar.header("📅 Select Despatch Date")
selected_range = st.sidebar.date_input("Despatch Date Range", [])
quick_range = st.sidebar.selectbox("🕒 Quick Despatch Range", [
    "None", "Yesterday", "Last 7 Days", "Last 30 Days", "Last 3 Months", "Last 6 Months", "Last 12 Months"
])

# TEMP LOAD to get available dates
@st.cache_data
def load_temp_dates():
    conn = connect_db()
    if conn:
        df = pd.read_sql("SELECT DISTINCT CAST(despatch_date AS DATE) AS despatch_date FROM OrdersDespatch", conn)
        conn.close()
        return sorted(pd.to_datetime(df['despatch_date']).unique())
    return []

available_dates = load_temp_dates()

if quick_range != "None":
    start_date, end_date = get_range_from_option(quick_range, available_dates)
elif len(selected_range) == 1:
    start_date = end_date = selected_range[0]
elif len(selected_range) == 2:
    start_date, end_date = selected_range
else:
    end_date = max(available_dates) if available_dates else datetime.today()
    start_date = end_date - timedelta(days=30)

start_date_str = start_date.strftime("%Y-%m-%d")
end_date_str = end_date.strftime("%Y-%m-%d")

# ------------------ LOAD DATA ------------------
@st.cache_data
def load_data(start_date_str, end_date_str):
    query = f"""
    WITH despatch_data AS (
        SELECT DISTINCT order_id, order_channel, despatch_date, order_value
        FROM OrdersDespatch
        WHERE CAST(despatch_date AS DATE) BETWEEN '{start_date_str}' AND '{end_date_str}'
    ),
    channel_total AS (
        SELECT 
            order_channel, 
            SUM(order_value) AS total_orders_value,
            COUNT(DISTINCT order_id) AS orders_count
        FROM despatch_data
        GROUP BY order_channel
    )
    SELECT order_channel AS channel, total_orders_value, orders_count
    FROM channel_total
    ORDER BY total_orders_value DESC;
    """
    conn = connect_db()
    if conn:
        df = pd.read_sql(query, conn)
        conn.close()
        return df
    return pd.DataFrame()

df = load_data(start_date_str, end_date_str)

if df.empty:
    st.warning("No orders found.")
    st.stop()

# ------------------ EXCEL EXPORT ------------------
grand_total_value = df["total_orders_value"].sum()
grand_total_count = df["orders_count"].sum()
df.loc[len(df.index)] = ["Grand Total", grand_total_value, grand_total_count]

output = io.BytesIO()
wb = Workbook()
ws = wb.active
ws.title = "Channel Summary"
ws["A1"] = "Selected Despatch Date:"
ws["B1"] = f"{start_date.strftime('%d-%m-%Y')} to {end_date.strftime('%d-%m-%Y')}"
ws["A2"] = "Day:"
ws["B2"] = start_date.strftime("%A") if start_date == end_date else "Multiple Days"
ws["A1"].font = ws["A2"].font = Font(bold=True)

for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 4):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 4 or row[0] == "Grand Total":
            cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 15
wb.save(output)
output.seek(0)

# ------------------ DISPLAY ------------------
st.subheader(f"📋 Channel Summary from {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
st.download_button("📅 Download Excel", data=output, file_name=f"Channel_Summary_{start_date_str}_to_{end_date_str}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
st.dataframe(df, use_container_width=True)

# ------------------ CHARTS ------------------
df_chart = df[df["channel"] != "Grand Total"]

st.subheader("📊 Total Orders Value by Channel")
st.plotly_chart(px.bar(df_chart, x="channel", y="total_orders_value", text="total_orders_value"), use_container_width=True)

st.subheader("📦 Orders Count by Channel")
st.plotly_chart(px.bar(df_chart, x="channel", y="orders_count", text="orders_count"), use_container_width=True)

st.subheader("🍩 Revenue Share by Channel")
st.plotly_chart(px.pie(df_chart, names='channel', values='total_orders_value', hole=0.4), use_container_width=True)

st.subheader("🍩 Orders Count Share by Channel")
st.plotly_chart(px.pie(df_chart, names='channel', values='orders_count', hole=0.4), use_container_width=True)
